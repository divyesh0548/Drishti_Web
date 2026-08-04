from pathlib import Path
from openpyxl import load_workbook


# Enter the Excel file path here
EXCEL_FILE_PATH = r"C:\Users\Divyesh Parmar\Downloads\V-8.xlsx"


def list_excel_sheets(file_path: str) -> list[str]:
    """Return all worksheet names from an Excel file."""

    excel_path = Path(file_path)

    if not excel_path.exists():
        raise FileNotFoundError(f"Excel file not found: {excel_path}")

    if excel_path.suffix.lower() not in {".xlsx", ".xlsm"}:
        raise ValueError("This program supports .xlsx and .xlsm files only.")

    workbook = load_workbook(
        filename=excel_path,
        read_only=True,
        data_only=True,
    )

    try:
        return workbook.sheetnames
    finally:
        workbook.close()


try:
    sheet_names = list_excel_sheets(EXCEL_FILE_PATH)

    print(f"Total sheets: {len(sheet_names)}")
    print("Sheet names:")

    for index, sheet_name in enumerate(sheet_names, start=1):
        print(f"{index}. {sheet_name}")

except (FileNotFoundError, ValueError, PermissionError) as error:
    print(f"Error: {error}")
except Exception as error:
    print(f"Unexpected error: {error}")